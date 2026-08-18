"""Build PenShowTracker.xlsx — the spreadsheet twin of the app.

    python3 tools/build_workbook.py      # writes penshow/PenShowTracker.xlsx

Row counts and output path can be overridden to produce a small copy for
testing, which is what tools/verify_workbook.py evaluates:

    PST_OUT=/tmp/penshow-verify/small.xlsx PST_INV=5 PST_SALE=6 PST_SHOW=2 \\
        python3 tools/build_workbook.py

Requires openpyxl.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

import os
OUT = os.environ.get("PST_OUT", "/home/user/ControlApp/penshow/PenShowTracker.xlsx")

INV_ROWS   = int(os.environ.get("PST_INV", 250))   # inventory data rows
SALE_ROWS  = int(os.environ.get("PST_SALE", 500))  # sales-log data rows
SHOW_ROWS  = int(os.environ.get("PST_SHOW", 20))   # show rows
HDR = 3               # header row on every data sheet

# Row bounds are needed before the sheets that reference them are built.
i0 = HDR + 1;  i1 = i0 + INV_ROWS - 1     # Inventory data rows
s0 = HDR + 1;  s1 = s0 + SALE_ROWS - 1    # Sales Log data rows
r0 = HDR + 1                              # Shows data rows

F = "Arial"
INK   = Font(name=F, size=10)
BOLD  = Font(name=F, size=10, bold=True)
TITLE = Font(name=F, size=16, bold=True, color="1F2937")
SUB   = Font(name=F, size=9, color="6B7280", italic=True)
HDRF  = Font(name=F, size=9, bold=True, color="FFFFFF")
INPUT = Font(name=F, size=10, color="0000FF")          # type here
CALC  = Font(name=F, size=10, color="000000")          # formula
LINK  = Font(name=F, size=10, color="008000")          # pulls from another sheet

HDRFILL  = PatternFill("solid", fgColor="1F2937")
BANDFILL = PatternFill("solid", fgColor="F3F4F6")
KEYFILL  = PatternFill("solid", fgColor="FFFF00")
GOODFILL = PatternFill("solid", fgColor="ECFDF5")
thin = Side(style="thin", color="D1D5DB")
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)

MONEY = '$#,##0.00;($#,##0.00);-'
MONEY0 = '$#,##0;($#,##0);-'
PCTF  = '0.0%'
INT   = '#,##0;(#,##0);-'

wb = openpyxl.Workbook()

# ----------------------------------------------------------------- helpers
def header(ws, title, subtitle, cols, widths, freeze):
    ws["A1"] = title; ws["A1"].font = TITLE
    ws["A2"] = subtitle; ws["A2"].font = SUB
    for i, (c, w) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=HDR, column=i, value=c)
        cell.font = HDRF; cell.fill = HDRFILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[HDR].height = 30
    ws.freeze_panes = freeze

def style_grid(ws, r0, r1, ncols, fonts, fmts):
    """fonts/fmts are per-column lists (1-indexed by position)."""
    for r in range(r0, r1 + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = fonts[c - 1]
            if fmts[c - 1]: cell.number_format = fmts[c - 1]
            cell.border = BOX
            if r % 2 == 0: cell.fill = BANDFILL

# ================================================================== LISTS
ls = wb.active; ls.title = "Lists"
ls["A1"] = "Validation lists"; ls["A1"].font = TITLE
ls["A2"] = "Add your own values down each column; the dropdowns pick them up automatically."
ls["A2"].font = SUB
BRAND_LIST = ["Submarine", "Swiss Brand", "Lamborghini", "NASA"]
TYPE_LIST  = ["Fountain Pen", "Ballpoint", "Rollerball", "Pen Set", "Ink", "Accessory"]
MAT_LIST   = ["Metal", "Acrylic", "Ebonite", "Wooden", "Resin"]
NIB_LIST   = ["EF", "F", "M", "B", "BB", "Stub", "Italic", "Music", "Flex", "n/a"]
LIST_COLS = {
    "A": ("Brand", BRAND_LIST),
    "B": ("Type", TYPE_LIST),
    "C": ("Nib", NIB_LIST),
    "F": ("Material", MAT_LIST),
    "D": ("Payment", ["Cash", "Zelle", "Venmo", "Card", "Check", "Trade", "Split"]),
    "E": ("Yes/No", ["Yes", "No"]),
}
for col, (name, vals) in LIST_COLS.items():
    ls[f"{col}{HDR}"] = name
    ls[f"{col}{HDR}"].font = HDRF; ls[f"{col}{HDR}"].fill = HDRFILL
    ls.column_dimensions[col].width = 18
    for i, v in enumerate(vals, start=HDR + 1):
        ls[f"{col}{i}"] = v; ls[f"{col}{i}"].font = INK

# ================================================================== SHOWS
sh = wb.create_sheet("Shows")
SHOW_COLS = ["Show", "City", "State", "Start", "End", "Sales tax rate", "Card fee %",
             "Card flat $", "Table / booth", "Travel", "Lodging", "Meals", "Shipping",
             "Other", "TOTAL SHOW COST", "Opening cash float", "Notes"]
header(sh, "Shows — the fixed costs of showing up",
       "Blue cells are yours to fill in. TOTAL SHOW COST is calculated. "
       "VERIFY every tax rate with the state before you sell — the seeded values are typical "
       "combined state+local rates for the venue city, not legal advice.",
       SHOW_COLS, [22, 22, 7, 11, 11, 12, 10, 10, 12, 11, 11, 11, 11, 11, 15, 14, 34],
       "C4")

SEED = [
    ("SF Pen Show",      "San Francisco Bay Area", "CA", 0.09375),
    ("Orlando Pen Show", "Orlando",                "FL", 0.065),
    ("Dallas Pen Show",  "Dallas",                 "TX", 0.0825),
    ("Denver Pen Show",  "Denver",                 "CO", 0.0881),
    ("Ohio Pen Show",    "Columbus",               "OH", 0.075),
]
r0 = HDR + 1
for i, (nm, city, st, rate) in enumerate(SEED):
    r = r0 + i
    sh.cell(row=r, column=1, value=nm)
    sh.cell(row=r, column=2, value=city)
    sh.cell(row=r, column=3, value=st)
    sh.cell(row=r, column=6, value=rate)
    sh.cell(row=r, column=7, value=0.026)
    sh.cell(row=r, column=8, value=0.10)
    sh.cell(row=r, column=6).fill = KEYFILL
for i in range(SHOW_ROWS):
    r = r0 + i
    sh.cell(row=r, column=15,
            value=f"=IFERROR(SUM(I{r}:N{r}),0)")
show_fonts = [INPUT]*14 + [CALC] + [INPUT, INPUT]
show_fmts  = [None, None, None, "yyyy-mm-dd", "yyyy-mm-dd", PCTF, PCTF, MONEY,
              MONEY0, MONEY0, MONEY0, MONEY0, MONEY0, MONEY0, MONEY0, MONEY0, None]
style_grid(sh, r0, r0 + SHOW_ROWS - 1, 17, show_fonts, show_fmts)
sh.cell(row=r0, column=6).comment = Comment(
    "Combined state + local rate as a decimal (9.375% = 0.09375).\n"
    "Confirm with the state's department of revenue and get a temporary "
    "seller's permit before the show. Colorado stacks home-rule district "
    "taxes on top of the state rate.", "Setup", width=320, height=120)

SHOW_RANGE = f"Shows!$A${r0}:$A${r0+SHOW_ROWS-1}"

# ============================================================== INVENTORY
iv = wb.create_sheet("Inventory")
INV_COLS = ["SKU", "Brand", "Type", "Material", "Model", "Colour / finish", "Nib",
            "Landed cost", "MRP", "Show price", "Floor price", "Dealer price",
            "Qty brought", "Qty sold", "Qty left", "Margin @ price", "Margin @ floor",
            "Profit / unit", "Cost tied up", "Retail value", "Case / tray",
            "Tags", "Photo link", "Notes"]
# Column letters are derived from the header list, so reordering a column can
# never silently point a formula at the wrong one.
IC = {name: get_column_letter(i) for i, name in enumerate(INV_COLS, start=1)}
header(iv, "Inventory — every pen you put on the table",
       "Blue = you type it. Black = calculated, do not overwrite. Qty sold reads the Sales Log "
       "automatically. Floor price is the lowest you will accept when someone haggles. "
       "Every pen is new, so there is no condition column.",
       INV_COLS, [18, 14, 15, 12, 18, 16, 8, 11, 10, 11, 11, 11, 9, 9, 9, 11, 11, 11, 11, 11,
                  14, 18, 26, 30], "F4")

for r in range(i0, i1 + 1):
    iv[f"{IC['Qty sold']}{r}"]      = (f'=IF($A{r}="","",IFERROR(SUMIFS(\'Sales Log\'!$E${s0}:$E${s1},'
                                       f'\'Sales Log\'!$C${s0}:$C${s1},$A{r}),0))')
    iv[f"{IC['Qty left']}{r}"]      = f'=IF($A{r}="","",{IC["Qty brought"]}{r}-{IC["Qty sold"]}{r})'
    iv[f"{IC['Margin @ price']}{r}"] = (f'=IF({IC["Show price"]}{r}>0,({IC["Show price"]}{r}-'
                                        f'{IC["Landed cost"]}{r})/{IC["Show price"]}{r},"")')
    iv[f"{IC['Margin @ floor']}{r}"] = (f'=IF({IC["Floor price"]}{r}>0,({IC["Floor price"]}{r}-'
                                        f'{IC["Landed cost"]}{r})/{IC["Floor price"]}{r},"")')
    iv[f"{IC['Profit / unit']}{r}"]  = f'=IF($A{r}="","",{IC["Show price"]}{r}-{IC["Landed cost"]}{r})'
    iv[f"{IC['Cost tied up']}{r}"]   = f'=IF($A{r}="","",{IC["Qty left"]}{r}*{IC["Landed cost"]}{r})'
    iv[f"{IC['Retail value']}{r}"]   = f'=IF($A{r}="","",{IC["Qty left"]}{r}*{IC["Show price"]}{r})'

inv_fonts = [INPUT]*13 + [LINK] + [CALC]*6 + [INPUT]*4
inv_fmts  = [None, None, None, None, None, None, None, MONEY, MONEY, MONEY, MONEY, MONEY,
             INT, INT, INT, PCTF, PCTF, MONEY, MONEY, MONEY, None, None, None, None]
style_grid(iv, i0, i1, len(INV_COLS), inv_fonts, inv_fmts)

# example row, clearly marked
for col, val in [("SKU", "SUB-FP-SHIKAR-M-BLK"), ("Brand", "Submarine"), ("Type", "Fountain Pen"),
                 ("Material", "Ebonite"), ("Model", "Shikari"),
                 ("Colour / finish", "Black"), ("Nib", "M"),
                 ("Landed cost", 6), ("MRP", 24), ("Show price", 18), ("Floor price", 14),
                 ("Dealer price", 12), ("Qty brought", 10), ("Case / tray", "Case 1"),
                 ("Notes", "EXAMPLE ROW — overwrite or delete it")]:
    iv[f"{IC[col]}{i0}"] = val
for c in range(1, len(INV_COLS) + 1):
    iv.cell(row=i0, column=c).fill = GOODFILL

for col, listcol, n in (("Brand", "A", len(BRAND_LIST)), ("Type", "B", len(TYPE_LIST)),
                        ("Nib", "C", len(NIB_LIST)), ("Material", "F", len(MAT_LIST))):
    dv = DataValidation(type="list",
                        formula1=f"=Lists!${listcol}${HDR+1}:${listcol}${HDR+n}",
                        allow_blank=True)
    iv.add_data_validation(dv)
    dv.add(f"{IC[col]}{i0}:{IC[col]}{i1}")

iv[f"{IC['SKU']}{i0}"].comment = Comment(
    "Nomenclature: BRAND-TYPE-MODEL[-NIB][-FINISH]\n"
    "  SUB-FP-SHIKAR-M-BLK   Submarine fountain pen, Shikari, medium, black\n"
    "  LAM-BP-AVENTA-CHR     Lamborghini ballpoint, Aventador, chrome\n"
    "Brands: SUB / SWB / LAM. Types: FP BP RB MP ST IN AC.\n"
    "The app generates these for you; leave blank on import and it will fill in.",
    "Setup", width=340, height=140)
iv[f"{IC['Floor price']}{i0}"].comment = Comment(
    "The lowest price you will take. Anything below this loses money once you "
    "count the table fee, so decide it now — not at 4pm on Sunday with a "
    "customer in front of you.", "Setup", width=300, height=100)

INV_SKU = f"Inventory!$A${i0}:$A${i1}"

# ============================================================== SALES LOG
sl = wb.create_sheet("Sales Log")
SALE_COLS = ["Date", "Show", "SKU", "Item (auto)", "Qty", "Unit price", "Discount $",
             "Unit cost (auto)", "Taxable", "Line revenue", "Sales tax", "Line COGS",
             "Line profit", "Payment method", "Payment ref", "Card fee", "Net to you",
             "Customer", "Email / phone", "Notes"]
header(sl, "Sales Log — one row per item sold",
       "Type SKU and the item name and cost fill themselves in. One row per line item: "
       "a customer buying three pens is three rows, same date and payment method.",
       SALE_COLS, [11, 20, 16, 26, 6, 11, 10, 11, 9, 12, 10, 11, 11, 14, 14, 10, 11,
                   18, 20, 26], "D4")

s0 = HDR + 1
s1 = s0 + SALE_ROWS - 1
for r in range(s0, s1 + 1):
    blank = f'$C{r}=""'
    # name + cost pulled from Inventory
    sl.cell(row=r, column=4, value=(
        f'=IF({blank},"",IFERROR(INDEX(Inventory!${IC["Brand"]}${i0}:${IC["Brand"]}${i1},MATCH($C{r},{INV_SKU},0))&" "&'
        f'INDEX(Inventory!${IC["Model"]}${i0}:${IC["Model"]}${i1},MATCH($C{r},{INV_SKU},0)),"SKU not in Inventory"))'))
    sl.cell(row=r, column=8, value=(
        f'=IF({blank},"",IFERROR(INDEX(Inventory!${IC["Landed cost"]}${i0}:${IC["Landed cost"]}${i1},'
        f'MATCH($C{r},{INV_SKU},0)),0))'))
    # revenue net of the line discount
    sl.cell(row=r, column=10, value=f'=IF({blank},"",MAX(0,$E{r}*$F{r}-$G{r}))')
    # tax uses the show's rate
    sl.cell(row=r, column=11, value=(
        f'=IF(OR({blank},$I{r}="No"),0,ROUND($J{r}*IFERROR(INDEX(Shows!$F${r0}:$F${r0+SHOW_ROWS-1},'
        f'MATCH($B{r},{SHOW_RANGE},0)),0),2))'))
    sl.cell(row=r, column=12, value=f'=IF({blank},"",$E{r}*$H{r})')
    sl.cell(row=r, column=13, value=f'=IF({blank},"",$J{r}-$L{r}-$P{r})')
    # card fee only on card payments
    sl.cell(row=r, column=16, value=(
        f'=IF(OR({blank},$N{r}<>"Card"),0,ROUND(($J{r}+$K{r})*'
        f'IFERROR(INDEX(Shows!$G${r0}:$G${r0+SHOW_ROWS-1},MATCH($B{r},{SHOW_RANGE},0)),0)+'
        f'IFERROR(INDEX(Shows!$H${r0}:$H${r0+SHOW_ROWS-1},MATCH($B{r},{SHOW_RANGE},0)),0),2))'))
    sl.cell(row=r, column=17, value=f'=IF({blank},"",$J{r}+$K{r}-$P{r})')

sale_fonts = [INPUT, INPUT, INPUT, LINK, INPUT, INPUT, INPUT, LINK, INPUT,
              CALC, CALC, CALC, CALC, INPUT, INPUT, CALC, CALC, INPUT, INPUT, INPUT]
sale_fmts  = ["yyyy-mm-dd", None, None, None, INT, MONEY, MONEY, MONEY, None,
              MONEY, MONEY, MONEY, MONEY, None, None, MONEY, MONEY, None, None, None]
style_grid(sl, s0, s1, 20, sale_fonts, sale_fmts)

dv_show = DataValidation(type="list", formula1=f"={SHOW_RANGE}", allow_blank=True)
sl.add_data_validation(dv_show); dv_show.add(f"B{s0}:B{s1}")
dv_sku = DataValidation(type="list", formula1=f"={INV_SKU}", allow_blank=True)
sl.add_data_validation(dv_sku); dv_sku.add(f"C{s0}:C{s1}")
dv_pay = DataValidation(type="list", formula1=f"=Lists!$D${HDR+1}:$D${HDR+7}", allow_blank=True)
sl.add_data_validation(dv_pay); dv_pay.add(f"N{s0}:N{s1}")
dv_yn = DataValidation(type="list", formula1=f"=Lists!$E${HDR+1}:$E${HDR+2}", allow_blank=True)
sl.add_data_validation(dv_yn); dv_yn.add(f"I{s0}:I{s1}")

sl.cell(row=s0, column=15).comment = Comment(
    "Zelle confirmation number, Venmo handle, or the last 4 of the card. "
    "This is what you match against your bank the same night — an unmatched "
    "transfer is unrecoverable once the buyer has walked away.", "Setup",
    width=300, height=100)

# =============================================================== DASHBOARD
db = wb.create_sheet("Dashboard")
DB_COLS = ["Show", "Sales lines", "Units", "Revenue", "COGS", "Gross profit", "Margin %",
           "Card fees", "Show costs", "NET PROFIT", "Break-even revenue",
           "Sales tax collected", "Cash", "Zelle", "Venmo", "Card", "Check", "Trade"]
header(db, "Dashboard — did the show actually make money?",
       "Everything here is calculated from the Shows and Sales Log tabs. "
       "Sales tax collected is a liability you remit to the state; it is never revenue and "
       "never profit.",
       DB_COLS, [22, 11, 8, 13, 13, 13, 10, 11, 12, 14, 15, 14, 12, 12, 12, 12, 12, 12], "B4")

d0 = HDR + 1
SL_SHOW = f"'Sales Log'!$B${s0}:$B${s1}"
for i in range(SHOW_ROWS):
    r = d0 + i
    src = f"Shows!$A${r0+i}"
    db.cell(row=r, column=1, value=f'=IF({src}="","",{src})')
    nm = f"$A{r}"
    db.cell(row=r, column=2,  value=f'=IF({nm}="","",COUNTIFS({SL_SHOW},{nm}))')
    db.cell(row=r, column=3,  value=f'=IF({nm}="","",SUMIFS(\'Sales Log\'!$E${s0}:$E${s1},{SL_SHOW},{nm}))')
    db.cell(row=r, column=4,  value=f'=IF({nm}="","",SUMIFS(\'Sales Log\'!$J${s0}:$J${s1},{SL_SHOW},{nm}))')
    db.cell(row=r, column=5,  value=f'=IF({nm}="","",SUMIFS(\'Sales Log\'!$L${s0}:$L${s1},{SL_SHOW},{nm}))')
    db.cell(row=r, column=6,  value=f'=IF({nm}="","",D{r}-E{r})')
    db.cell(row=r, column=7,  value=f'=IF(N(D{r})>0,F{r}/D{r},"")')
    db.cell(row=r, column=8,  value=f'=IF({nm}="","",SUMIFS(\'Sales Log\'!$P${s0}:$P${s1},{SL_SHOW},{nm}))')
    db.cell(row=r, column=9,  value=f'=IF({nm}="","",IFERROR(INDEX(Shows!$O${r0}:$O${r0+SHOW_ROWS-1},MATCH({nm},{SHOW_RANGE},0)),0))')
    db.cell(row=r, column=10, value=f'=IF({nm}="","",F{r}-H{r}-I{r})')
    db.cell(row=r, column=11, value=f'=IF(N(G{r})>0,(I{r}+H{r})/G{r},"")')
    db.cell(row=r, column=12, value=f'=IF({nm}="","",SUMIFS(\'Sales Log\'!$K${s0}:$K${s1},{SL_SHOW},{nm}))')
    for j, m in enumerate(["Cash", "Zelle", "Venmo", "Card", "Check", "Trade"]):
        db.cell(row=r, column=13 + j, value=(
            f'=IF({nm}="","",SUMIFS(\'Sales Log\'!$Q${s0}:$Q${s1},{SL_SHOW},{nm},\'Sales Log\'!$N${s0}:$N${s1},"{m}"))'))

db_fonts = [LINK] + [CALC]*17
db_fmts  = [None, INT, INT, MONEY0, MONEY0, MONEY0, PCTF, MONEY0, MONEY0, MONEY0,
            MONEY0, MONEY0, MONEY0, MONEY0, MONEY0, MONEY0, MONEY0, MONEY0]
style_grid(db, d0, d0 + SHOW_ROWS - 1, 18, db_fonts, db_fmts)

tr = d0 + SHOW_ROWS
db.cell(row=tr, column=1, value="SEASON TOTAL")
for c in list(range(2, 7)) + list(range(8, 11)) + list(range(12, 19)):
    L = get_column_letter(c)
    db.cell(row=tr, column=c, value=f"=SUM({L}{d0}:{L}{d0+SHOW_ROWS-1})")
db.cell(row=tr, column=7, value=f'=IF(N(D{tr})>0,F{tr}/D{tr},"")')
db.cell(row=tr, column=11, value=f'=IF(N(G{tr})>0,(I{tr}+H{tr})/G{tr},"")')
for c in range(1, 19):
    cell = db.cell(row=tr, column=c)
    cell.font = Font(name=F, size=11, bold=True)
    cell.fill = PatternFill("solid", fgColor="DBEAFE")
    cell.border = BOX
    if c > 1: cell.number_format = db_fmts[c - 1]

# inventory position block
p = tr + 3
db.cell(row=p, column=1, value="Inventory position").font = Font(name=F, size=12, bold=True)
blocks = [
    ("SKUs listed",            f'=COUNTA(Inventory!$A${i0}:$A${i1})', INT),
    ("Units brought",          f'=SUM(Inventory!${IC["Qty brought"]}${i0}:${IC["Qty brought"]}${i1})', INT),
    ("Units sold",             f'=SUM(Inventory!${IC["Qty sold"]}${i0}:${IC["Qty sold"]}${i1})', INT),
    ("Units left",             f'=SUM(Inventory!${IC["Qty left"]}${i0}:${IC["Qty left"]}${i1})', INT),
    ("Cost still on the table", f'=SUMIF(Inventory!${IC["Qty left"]}${i0}:${IC["Qty left"]}${i1},">0",'
                               f'Inventory!${IC["Cost tied up"]}${i0}:${IC["Cost tied up"]}${i1})', MONEY0),
    ("Retail still on the table", f'=SUMIF(Inventory!${IC["Qty left"]}${i0}:${IC["Qty left"]}${i1},">0",'
                                 f'Inventory!${IC["Retail value"]}${i0}:${IC["Retail value"]}${i1})', MONEY0),
    ("Sell-through rate",      f'=IFERROR(SUM(Inventory!${IC["Qty sold"]}${i0}:${IC["Qty sold"]}${i1})/'
                               f'SUM(Inventory!${IC["Qty brought"]}${i0}:${IC["Qty brought"]}${i1}),"")', PCTF),
]
for i, (label, f_, fmt) in enumerate(blocks):
    rr = p + 1 + i
    db.cell(row=rr, column=1, value=label).font = INK
    c = db.cell(row=rr, column=2, value=f_)
    c.font = CALC; c.number_format = fmt
    db.cell(row=rr, column=1).border = BOX; c.border = BOX

db.cell(row=p, column=4, value="Read this before the next show").font = Font(name=F, size=12, bold=True)
notes = [
    "NET PROFIT is what you keep after cost of goods, card fees, table, travel, hotel and meals.",
    "Break-even revenue is what you must sell, at your current margin, just to cover the show.",
    "Sales tax collected is owed to the state. Set it aside; it is not income.",
    "Cash / Zelle / Venmo / Card columns are your end-of-day reconciliation — count against them nightly.",
    "Sell-through under ~40% usually means you brought too much, or priced too high.",
]
for i, n in enumerate(notes):
    c = db.cell(row=p + 1 + i, column=4, value="• " + n)
    c.font = Font(name=F, size=9, color="374151")
    c.alignment = Alignment(wrap_text=False)

# =============================================================== READ ME
rm = wb.create_sheet("Start here")
rm.column_dimensions["A"].width = 4
rm.column_dimensions["B"].width = 112
rm["B2"] = "Pen Show Tracker"; rm["B2"].font = Font(name=F, size=20, bold=True)
rm["B3"] = "Submarine · Swiss Brand · Lamborghini — inventory, pricing, sales and P&L for the 2026 US pen show circuit."
rm["B3"].font = SUB

LINES = [
    ("h", "How to use it"),
    ("n", "1.  Shows tab — fill in your table fee, travel, lodging and meals for each show, and VERIFY the sales tax rate."),
    ("n", "2.  Inventory tab — one row per pen. Pick brand and type from the dropdowns, then type cost, price and floor."),
    ("n", "3.  Sales Log tab — one row per item sold. Pick the SKU from the dropdown; name, cost and tax fill in for you."),
    ("n", "4.  Dashboard tab — per-show P&L and your end-of-day payment reconciliation. Nothing to type here."),
    ("", ""),
    ("h", "Colour code"),
    ("i", "Blue text        you type it"),
    ("c", "Black text       calculated — overwriting it breaks the sheet"),
    ("l", "Green text       pulled from another tab"),
    ("k", "Yellow fill      a key assumption to confirm before the show"),
    ("", ""),
    ("h", "What this sheet does NOT do — use the phone app in the same folder for these"),
    ("n", "•  Work as a till while a customer is standing in front of you. Typing rows into a spreadsheet on a phone at a"),
    ("n", "   crowded table is where the errors come from. The app is a two-tap point of sale; export from it into here."),
    ("n", "•  Hold stock for a customer coming back on Sunday."),
    ("n", "•  Store photos of each pen."),
    ("n", "•  Warn you when you are about to sell below your floor price."),
    ("", ""),
    ("h", "Before the first show — the things that bite people"),
    ("n", "•  Sales tax: each state wants its own temporary seller's permit BEFORE you take a dollar. Colorado is the"),
    ("n", "   hardest — home-rule cities stack their own tax on top of the state rate."),
    ("n", "•  Zelle: most banks cap sends at $500–$2,500 a day. A single high-end pen will bounce off that limit at the"),
    ("n", "   table. Raise the limit in advance and know your backup method. Zelle is also irreversible — no chargebacks,"),
    ("n", "   which protects you, but also means a mistyped amount is gone."),
    ("n", "•  Venmo: use a business profile. Taking business payments on a personal profile violates the terms and can"),
    ("n", "   get funds frozen mid-show. Business profiles charge a seller fee — put it in the Card fee column if you use one."),
    ("n", "•  1099-K: Venmo and card processors report your volume. Keep the Sales Log clean and this is a non-event."),
    ("n", "•  Landed cost means the pen plus shipping plus any import fee. Put the real number in or the P&L is fiction."),
    ("n", "•  Insurance: your homeowner's policy almost certainly does not cover inventory in a hotel room or in transit."),
    ("n", "•  Back this file up every night. One file on one laptop is not a backup."),
]
r = 6
for kind, text in LINES:
    c = rm.cell(row=r, column=2, value=text)
    if kind == "h":
        c.font = Font(name=F, size=12, bold=True, color="1F2937")
        rm.row_dimensions[r].height = 24
    elif kind == "i": c.font = INPUT
    elif kind == "c": c.font = CALC
    elif kind == "l": c.font = LINK
    elif kind == "k": c.font = INK; c.fill = KEYFILL
    else: c.font = INK
    r += 1

# Order the tabs the way they are actually used, and tuck the lookup lists away.
ORDER = ["Start here", "Dashboard", "Inventory", "Sales Log", "Shows", "Lists"]
wb._sheets = [wb[n] for n in ORDER]
wb["Lists"].sheet_state = "hidden"
wb.active = 0

for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False

wb.save(OUT)
print("wrote", OUT)
