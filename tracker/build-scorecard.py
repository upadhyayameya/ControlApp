#!/usr/bin/env python3
"""
Build the weekly company scorecard.

This is Harry's scorecard, rebuilt so the colours happen on their own. Every
row, goal and section is taken from the workbook he has been sending out --
nothing is renamed and nothing is invented, so it stays the sheet the team
already recognises.

WHAT CHANGED FROM THE SAMPLE

  * The week-ending columns are generated for the month rather than typed.
    Harry's sheet hard-codes four; October 2026 has five Fridays, and a fifth
    week with nowhere to go is how a month quietly loses a week of numbers.

  * Cells colour themselves. Green / Yellow / Red against the goal, as a real
    conditional format, so it keeps working when somebody types into it next
    month without opening this script.

  * A Source column says where each row's number comes from. Four rows can be
    filled from monday.com with no judgement involved; the rest are finance
    and sales figures that live nowhere but in Harry's head and his invoices.
    Marking which is which is the difference between a number you can trust
    and a number you have to go and check.

WHAT IS NOT AUTO-FILLED, AND WHY

  The per-utility pre-approval dollars look automatable and are not. The KPI
  board carries no utility column, so the only route is matching a KPI row's
  name to a project's name -- which lands 84% of rows, and monday's utility
  values ("BGE") cannot tell Harry's four separate BGE programme lines apart.
  An 84%-complete money figure that management acts on is worse than a blank
  one they know to fill.
"""
import argparse, datetime, json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# Met the goal / within striking distance / behind. One place to change them.
YELLOW_AT = 0.80

GREEN  = PatternFill('solid', start_color='C6EFCE', end_color='C6EFCE')
YELLOW = PatternFill('solid', start_color='FFEB9C', end_color='FFEB9C')
RED    = PatternFill('solid', start_color='FFC7CE', end_color='FFC7CE')
HEADFILL = PatternFill('solid', start_color='1F3864', end_color='1F3864')
SUBFILL  = PatternFill('solid', start_color='D9E1F2', end_color='D9E1F2')
AUTOFILL = PatternFill('solid', start_color='EDF3FF', end_color='EDF3FF')

HEAD = Font(bold=True, color='FFFFFF')
BOLD = Font(bold=True)
SMALL = Font(size=9, italic=True, color='595959')
THIN = Side(style='thin', color='BFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

REVENUE = [
    ('Tune-Up Invoiced Billed', 800000),
    ('Commissioning Billed', 50000),
    ('Consulting / Ongoing Services Billed', 45000),
    ('BAS / Controls Services Billed', 105000),
    ('Direct Install (HVAC / Misc.)', 600000),
]
PA_SUBMISSIONS = [
    ('BGE - BPTU - Phase 1 Incentive', 86700),
    ('BGE - BPTU - Phase 2 Incentive', 300000),
    ('BGE - Tune-Up', 200000),
    ('BGE - HVAC Tune-Up', 50000),
    ('Pepco / WGL', 180000),
    ('Delmarva', 40000),
    ('SMECO', 5000),
    ('GA', 49000),
    ('CGS (VA)', 60000),
]
# No goals on these in Harry's sheet -- they are volume, not money.
COUNTERS = [
    'Pre-Approvals Assigned',
    'Pre-Approvals Complete',
    '% Completion Rate',
    'SOW Sent to Customer',
    'Implementations Scheduled',
    'Closeouts Submitted',
    'Closeouts Received',
]
SALES = [
    ('BGE - BPTU - Phase 1 Incentive', 86700),
    ('BGE - BPTU - Phase 2 Incentive', 300000),
    ('BGE - Tune-Up', 200000),
    ('HVAC TUNEUP PRESCRIPTIVE', 50000),
    ('Pepco', 240000),
    ('Delmarva', 60000),
    ('SMECO', 5000),
    ('WGL', 2500),
    ('GA - CGS', 49000),
    ('VA - CGS', 25000),
    ('Direct Install', 500000),
    ('BAS (Installs / Service Work)', 100000),
    ('Commissioning', 25000),
    ('Ongoing BEPS Contracts', 50000),
    ('HVAC Service Contracts', 2500),
    ('BAS Service Contracts', None),
    ('Zentility (# Contracts Signed)', None),
]
# Counter rows monday can fill with no judgement: the KPI board's own groups,
# and the SOW date off the TU tracker. Names match Harry's rows exactly.
AUTO_ROWS = {'SOW Sent to Customer', 'Implementations Scheduled',
             'Closeouts Submitted', 'Closeouts Received'}


def week_endings(year, month):
    """Every Friday in the month -- four some months, five others."""
    d = datetime.date(year, month, 1)
    nxt = datetime.date(year + (month == 12), (month % 12) + 1, 1)
    out = []
    while d < nxt:
        if d.weekday() == 4:
            out.append(d)
        d += datetime.timedelta(days=1)
    return out


def build(year, month, auto, out_path):
    weeks = week_endings(year, month)
    n = len(weeks)
    first = 2                      # column B
    last = first + n - 1
    total_col = last + 1
    goal_col = total_col + 1
    src_col = goal_col + 1
    L = get_column_letter
    month_name = datetime.date(year, month, 1).strftime('%B %Y')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Scorecard'
    r = 1
    ws.cell(r, 1, 'HBS WEEKLY SCORECARD').font = Font(bold=True, size=14, color='1F3864')
    ws.cell(r, 2, f'Week Ending - {month_name}').font = Font(bold=True, size=12)
    r += 1
    ws.cell(r, 1, f'Green = goal met  ·  Yellow = within {int(YELLOW_AT*100)}%  ·  '
                  f'Red = below {int(YELLOW_AT*100)}%.  Weekly cells score against the '
                  f'monthly goal split over {n} weeks; the Total column scores against the whole goal.'
            ).font = SMALL
    r += 2

    def section(title, rows, start_row):
        """One block: header, its rows, and a team total."""
        rr = start_row
        ws.cell(rr, 1, title).font = HEAD
        ws.cell(rr, 1).fill = HEADFILL
        for i, w in enumerate(weeks):
            c = ws.cell(rr, first + i, w)
            c.number_format = 'm/d'
            c.font = HEAD; c.fill = HEADFILL; c.alignment = Alignment(horizontal='center')
        for col, lab in ((total_col, 'Total'), (goal_col, 'Monthly Goal'), (src_col, 'Source')):
            c = ws.cell(rr, col, lab); c.font = HEAD; c.fill = HEADFILL
            c.alignment = Alignment(horizontal='center')
        rr += 1
        body_start = rr
        for label, goal in rows:
            ws.cell(rr, 1, label).border = BOX
            filled = auto.get(label) if label in AUTO_ROWS else None
            for i, w in enumerate(weeks):
                c = ws.cell(rr, first + i)
                if filled is not None:
                    c.value = filled.get(w.isoformat())
                    c.fill = AUTOFILL
                c.border = BOX
                if goal is not None:
                    c.number_format = '#,##0'
            t = ws.cell(rr, total_col, f'=SUM({L(first)}{rr}:{L(last)}{rr})')
            t.font = BOLD; t.border = BOX
            if goal is not None:
                t.number_format = '#,##0'
                g = ws.cell(rr, goal_col, goal)
                g.number_format = '#,##0'; g.border = BOX
            else:
                ws.cell(rr, goal_col, '').border = BOX
            ws.cell(rr, src_col, 'monday.com' if filled is not None else 'manual').font = SMALL
            rr += 1
        body_end = rr - 1
        rr += 1
        ws.cell(rr, 1, 'Total for Team').font = BOLD
        for col in list(range(first, last + 1)) + [total_col, goal_col]:
            c = ws.cell(rr, col, f'=SUM({L(col)}{body_start}:{L(col)}{body_end})')
            c.font = BOLD; c.number_format = '#,##0'; c.fill = SUBFILL; c.border = BOX
        return body_start, body_end, rr + 2

    def rag(body_start, body_end):
        """Colour the weekly cells against a week's share, the Total against the month.

        Every rule is guarded on the goal being present and the cell holding a
        real number, so an empty sheet at the start of a month stays white
        instead of turning solid red and teaching everyone to ignore it."""
        wk = f'{L(first)}{body_start}:{L(last)}{body_end}'
        tot = f'{L(total_col)}{body_start}:{L(total_col)}{body_end}'
        g = f'${L(goal_col)}{body_start}'
        cell = f'{L(first)}{body_start}'
        for rng, target, ref in ((wk, f'{g}/{n}', cell), (tot, g, f'{L(total_col)}{body_start}')):
            live = f'AND({g}<>"",ISNUMBER({ref}),{ref}<>0)'
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({live},{ref}>={target})'], fill=GREEN, stopIfTrue=True))
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({live},{ref}>={YELLOW_AT}*{target})'], fill=YELLOW, stopIfTrue=True))
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({live},{ref}<{YELLOW_AT}*{target})'], fill=RED, stopIfTrue=True))

    s, e, r = section('Company Revenue', REVENUE, r); rag(s, e)
    s, e, r = section('Pre-Approval Submissions', PA_SUBMISSIONS, r); rag(s, e)

    # The counter block has no goals, so it is listed rather than scored.
    ws.cell(r, 1, 'Pipeline Volume').font = HEAD
    ws.cell(r, 1).fill = HEADFILL
    for i, w in enumerate(weeks):
        c = ws.cell(r, first + i, w); c.number_format = 'm/d'
        c.font = HEAD; c.fill = HEADFILL; c.alignment = Alignment(horizontal='center')
    for col, lab in ((total_col, 'Total'), (goal_col, ''), (src_col, 'Source')):
        c = ws.cell(r, col, lab); c.font = HEAD; c.fill = HEADFILL
    r += 1
    pct_row = None
    for label in COUNTERS:
        ws.cell(r, 1, label).border = BOX
        if label == '% Completion Rate':
            pct_row = r
            for i in range(n):
                c = ws.cell(r, first + i,
                            f'=IFERROR({L(first+i)}{r-1}/{L(first+i)}{r-2},"")')
                c.number_format = '0%'; c.border = BOX
            t = ws.cell(r, total_col,
                        f'=IFERROR({L(total_col)}{r-1}/{L(total_col)}{r-2},"")')
            t.number_format = '0%'; t.font = BOLD; t.border = BOX
            ws.cell(r, src_col, 'calculated').font = SMALL
        else:
            filled = auto.get(label) if label in AUTO_ROWS else None
            for i, w in enumerate(weeks):
                c = ws.cell(r, first + i)
                if filled is not None:
                    c.value = filled.get(w.isoformat()); c.fill = AUTOFILL
                c.border = BOX
            t = ws.cell(r, total_col, f'=SUM({L(first)}{r}:{L(last)}{r})')
            t.font = BOLD; t.border = BOX
            ws.cell(r, src_col, 'monday.com' if filled is not None else 'manual').font = SMALL
        r += 1
    r += 1

    s, e, r = section('Sales (Pre-Approvals / New Work)', SALES, r); rag(s, e)

    ws.cell(r, 1, 'Rows shaded pale blue are filled automatically from monday.com. '
                  'Everything else is typed in — those figures are not on any board.').font = SMALL
    ws.column_dimensions['A'].width = 42
    for i in range(n):
        ws.column_dimensions[L(first + i)].width = 13
    ws.column_dimensions[L(total_col)].width = 13
    ws.column_dimensions[L(goal_col)].width = 14
    ws.column_dimensions[L(src_col)].width = 12
    ws.freeze_panes = 'B5'
    wb.save(out_path)
    return weeks


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--month', required=True, help='YYYY-MM')
    ap.add_argument('--auto', help='JSON of {row label: {week-ending ISO: value}}')
    ap.add_argument('--out', required=True)
    a = ap.parse_args()
    y, m = (int(x) for x in a.month.split('-'))
    auto = json.load(open(a.auto)) if a.auto else {}
    wks = build(y, m, auto, a.out)
    print(f'wrote {a.out}: {len(wks)} week-endings '
          f'({", ".join(w.isoformat() for w in wks)})')
